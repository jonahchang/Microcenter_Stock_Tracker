import re
import time
import os
import sys
import json
import tkinter as tk
from tkinter import Tk, filedialog, simpledialog, messagebox
import undetected_chromedriver as uc
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import range_boundaries
import random

DEBUG_MODE = True

json_file = "MC_products.json"
original_file = "MC_original_products.json"

def load_products():
    global power_supplies, coolers, chassis, miscellaneous

    # Load product list
    if os.path.exists(json_file):
        with open(json_file, "r") as f:
            data = json.load(f)
            power_supplies = data.get("power_supplies", {})
            coolers = data.get("coolers", {})
            chassis = data.get("chassis", {})
            miscellaneous = data.get("miscellaneous", {})
    else:
        # Power Supplies
        power_supplies = {"ROG-THOR-1000P2-GAMING": "https://www.microcenter.com/product/666611/asus-rog-thor-1000-watt-80-plus-platinum-atx-fully-modular-power-supply",
                        "ROG-STRIX-1200P-GAMING": "https://www.microcenter.com/product/695232/asus-rog-strix-1200-watt-80-plus-platinum-atx-fully-modular-power-supply-atx-31-compatible",
                        "ROG-LOKI-1000P-SFX-L-GAMING": "https://www.microcenter.com/product/664884/asus-rog-loki-1000-watt-80-plus-platinum-sfx-l-fully-modular-power-supply-black-atx-30-compatible",
                        "ROG-LOKI-850P-WHITE-SFX-L-GAMING": "https://www.microcenter.com/product/664885/asus-rog-loki-850-watt-80-plus-platinum-sfx-l-fully-modular-power-supply-white-atx-30-compatible",
                        "ROG-LOKI-850P-SFX-L-GAMING": "https://www.microcenter.com/product/664883/asus-rog-loki-850-watt-80-plus-gold-sfx-l-fully-modular-power-supply-black-atx-30-compatible",
                        "ROG-STRIX-850G-AURA-GAMING": "https://www.microcenter.com/product/669273/asus-rog-strix-gold-aura-edition-850-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "ROG-STRIX-750G-AURA-GAMING": "https://www.microcenter.com/product/669274/asus-rog-strix-gold-aura-edition-750-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "TUF-GAMING-1200G": "https://www.microcenter.com/product/676964/asus-tuf-gaming-1200-watt-80-plus-gold-atx-fully-modular-power-supply",
                        "TUF-GAMING-1000G": "https://www.microcenter.com/product/665308/asus-tuf-gaming-1000-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "TUF-GAMING-850G": "https://www.microcenter.com/product/665319/asus-tuf-gaming-850-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "TUF-GAMING-750G": "https://www.microcenter.com/product/665320/asus-tuf-gaming-750-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "AP-850G": "https://www.microcenter.com/product/675843/asus-prime-850-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible",
                        "AP-750G": "https://www.microcenter.com/product/675842/asus-prime-750-watt-80-plus-gold-atx-fully-modular-power-supply-atx-30-compatible"}
        
        # Coolers
        coolers = {"ROG RYUJIN III 360 ARGB EXTREME WHT": "https://www.microcenter.com/product/690065/asus-asus-rog-ryuyjin-iii-360-argb-extreme-360mm-all-in-one-liquid-cpu-cooling-kit-white",
                    "ROG RYUJIN III 360 ARGB EXTREME": "https://www.microcenter.com/product/690066/asus-rog-ryujin-iii-argb-extreme-360mm-all-in-one-liquid-cpu-cooling-kit-black",
                    "ROG RYUJIN III 360": "https://www.microcenter.com/product/668461/asus-rog-ryujin-iii-360mm-all-in-one-liquid-cpu-cooling-kit",
                    "ProArt LC 420": "https://www.microcenter.com/product/678856/asus-proart-lc-420mm-all-in-one-liquid-cpu-cooling-kit-black"}

        # Chassis
        chassis = {"GR701 ROG HYPERION": "https://www.microcenter.com/product/664435/asus-asus-rog-hyperion-gr701-tempered-glass-eatx-full-tower-computer-case-black",
                    "GX601 ROG STRIX HELIOS CASE/WT/AL/WITH HANDLE": "https://www.microcenter.com/product/625183/asus-rog-strix-helios-gx601-rgb-tempered-glass-atx-mid-tower-computer-case-white-edition",
                    "GX601 ROG STRIX HELIOS CASE/BK/AL/WITH HANDLE": "https://www.microcenter.com/product/609942/asus-rog-strix-helios-gx601-rgb-tempered-glass-atx-mid-tower-computer-case-black",
                    "PA602 ProArt Case": "https://www.microcenter.com/product/676302/asus-proart-pa602-tempered-glass-eatx-mid-tower-computer-case-black",
                    "PROART PA401 WOOD TG PWM BLACK": "https://www.microcenter.com/product/690056/asus-proart-pa401-wood-edition-tempered-glass-atx-mid-tower-computer-case-black",
                    "GT502 TUF GAMING CASE/BLK": "https://www.microcenter.com/product/662252/asus-tuf-gaming-gt502-tempered-glass-atx-mid-tower-computer-case-black",
                    "GT502 TUF GAMING CASE/WHT": "https://www.microcenter.com/product/662254/asus-tuf-gaming-gt502-tempered-glass-atx-mid-tower-computer-case-white",
                    "GT501 TUF GAMING CASE/GRY/WITH HANDLE": "https://www.microcenter.com/product/601243/asus-tuf-gaming-gt501-rgb-tempered-glass-atx-mid-tower-computer-case",
                    "TUF GAMING GT302 ARGB BLACK": "https://www.microcenter.com/product/679946/asus-tuf-gaming-gt302-argb-tempered-glass-atx-mid-tower-computer-case-black",
                    "TUF GAMING GT302 ARGB WHT": "https://www.microcenter.com/product/679945/asus-tuf-gaming-gt302-argb-tempered-glass-atx-mid-tower-computer-case-white",
                    "A31 PLUS/BK/TG/ARGB// ": "https://www.microcenter.com/product/690543/asus-a31-plus-tempered-glass-atx-mid-tower-computer-case-black",
                    "AP201 ASUS PRIME CASE MESH": "https://www.microcenter.com/product/651914/asus-prime-ap201-microatx-mini-tower-computer-case-black",
                    "AP201 ASUS PRIME CASE MESH WHITE EDITION": "https://www.microcenter.com/product/651917/asus-prime-ap201-microatx-mini-tower-computer-case-white"}
        
        # Miscellaneous
        miscellaneous = {}

        save_products(original_file)

        save_products(json_file)

def save_products(filename):
    with open(filename, "w") as f:
        json.dump({
            "power_supplies": power_supplies,
            "coolers": coolers,
            "chassis": chassis,
            "miscellaneous": miscellaneous
        }, f, indent=2)

# Mapping of store ID to store name (must match Excel header format)
store_map = {
    "195": "Santa Clara", "101": "Tustin", "181": "Denver", "185": "Miami", "065": "Duluth", "041": "Marietta", 
    "151": "Chicago", "025": "Westmont", "165": "Indianapolis", "191": "Overland Park", "121": "Cambridge", 
    "085": "Rockville", "125": "Parkville", "055": "Madison Heights", "045": "St Louis Park", "095": "Brentwood", 
    "175": "Charlotte", "075": "New Jersey", "171": "Westbury", "115": "Brooklyn", "145": "Flushing", "105": "Yonkers", 
    "141": "Colombus", "051": "Mayfield Heights", "071": "Sharonville", "061": "St Davids", "155": "Houston",
    "131": "Dallas", "081": "Fairfax"
}

# Stock highlighting colors
green_fill = PatternFill(start_color='00FF00', end_color='00FF00',  fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00',  fill_type='solid')
blue_fill = PatternFill(start_color='83CCEB', end_color='83CCEB',  fill_type='solid')

def process_add_products_sheet(wb):
    if "Add Products" not in wb.sheetnames:
        return
    
    sheet = wb["Add Products"]
    row = 3
    added_any = False

    while True:
        name_cell = sheet[f"A{row}"]
        url_cell = sheet[f"B{row}"]
        cat_cell = sheet[f"C{row}"]

        name = name_cell.value.strip() if name_cell.value else ""
        url = url_cell.value.strip() if url_cell.value else ""
        category = (cat_cell.value or "").strip().lower()

        if not name and not url and not category:
            break

        if not name or not url:
            row += 1
            continue

        if not category:
            category = "miscellaneous"

        category_map = {
            "power supply": power_supplies,
            "cooler": coolers,
            "chassis": chassis,
            "miscellaneous": miscellaneous
        }

        target_dict = category_map.get(category.lower(), miscellaneous)
        if name not in target_dict:
            target_dict[name] = url
            added_any = True

        for col in ["A", "B", "C"]:
            sheet[f"{col}{row}"].value = None

        row += 1

    rows = list(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3))
    clean_data = [[cell.value for cell in row] for row in rows if any(cell.value for cell in row)]

    for i in range(3, sheet.max_row + 1):
        for col in ["A", "B", "C"]:
            sheet[f"{col}{i}"].value = None

    for idx, (name, url, category) in enumerate(clean_data, start=3):
        sheet[f"A{idx}"] = name
        sheet[f"B{idx}"] = url
        sheet[f"C{idx}"] = category

    if added_any:
        save_products(json_file)

def normalize_cell_value(value):
    if value is None:
        return 0
    val = str(value).strip().upper()
    if val == 'NO STOCK':
        return 0
    if val == 'N/A':
        return 0
    if val == '25+':
        return 25
    try:
        int_val = int(val)
        return 25 if int_val >= 25 else int_val
    except ValueError:
        return 0
    
def analyze_stock(wb):
    sheetnames = wb.sheetnames
    
    for i in range(3, len(sheetnames)):
        prev_sheet = wb[sheetnames[i - 1]]
        curr_sheet = wb[sheetnames[i]]

        def sheet_to_dict(sheet):
            data = {}
            for row in sheet.iter_rows(min_row=2):
                name = str(row[1].value).strip() if row[1].value else None
                if not name:
                    continue
                values = [normalize_cell_value(cell.value) for cell in row[2:31]]
                data[name] = values
            return data
        
        prev_data = sheet_to_dict(prev_sheet)

        for row in curr_sheet.iter_rows(min_row=2):
            product_name = str(row[1].value).strip() if row[1].value else None
            if not product_name or product_name not in prev_data:
                continue

            prev_row = prev_data[product_name]
            curr_row = [normalize_cell_value(cell.value) for cell in row[2:31]]

            for j, (prev_val, curr_val) in enumerate(zip(prev_row, curr_row), start=2):
                cell = row[j]
                if(prev_val == 'N/A' and isinstance(curr_val, int)) or (curr_val == 'N/A' and isinstance(prev_val, int)):
                    cell.fill = blue_fill
                elif isinstance(prev_val, int) and isinstance(curr_val, int):
                    if curr_val > prev_val:
                        cell.fill = green_fill
                    elif curr_val < prev_val:
                        cell.fill = yellow_fill

# Format the sheet
def format_new_sheet(ws):
    category_positions = {}
    row_start = 2
    for category, products in [("Power Supply", power_supplies),
                               ("AIO Liquid CPU Cooler", coolers),
                               ("Chassis", chassis),
                               ("Miscellaneous", miscellaneous)]:
        if not products:
            continue
        row_end = row_start + len(products) - 1
        ws.merge_cells(start_row=row_start, start_column=1, end_row = row_end, end_column=1)
        ws[f"A{row_start}"] = category
        ws[f"AH{row_start}"] = category.upper()
        category_positions[category] = row_start
        for i, name in enumerate(products.keys(), start=row_start):
            ws[f"B{i}"] = name
        row_start = row_end + 1

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=31):
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    return category_positions

def product_sums(ws, category_positions):
    def get_merged_value(ws, cell_ref):
        cell = ws[cell_ref]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws[merged_range.coord.split(":")[0]].value
        return cell.value

    for row in range(2, ws.max_row + 1):
        ws[f"AF{row}"] = f"=SUM(C{row}:AE{row})"
        ws[f"AH{row}"] = f"=COUNTIF(C{row}:AE{row}, 0)"

    category_blocks = {}
    current_category = None
    current_start = None

    for row in range(2, ws.max_row + 1):
        category_value = get_merged_value(ws, f"A{row}")

        if category_value != current_category:
            # Close the previous category block
            if current_category is not None:
                category_blocks[current_category] = (current_start, row)
            # Start a new category block
            current_category = category_value
            current_start = row

    # Close the last category block
    if current_category is not None:
        category_blocks[current_category] = (current_start, ws.max_row + 1)

    # 3. Write category totals in AG at the top row of each category
    for category, (start_row, end_row) in category_blocks.items():
        total_rows = []
        for row in range(start_row, end_row):
            if not get_merged_value(ws, f"A{row}"):  # Stop if blank
                break
            total_rows.append(f"AF{row}")

        if total_rows:
            ws[f"AG{start_row}"] = f"=SUM({','.join(total_rows)})"

def prepare_chart_data(wb):
    # Ensure Charts sheet exists in position 3
    if len(wb.sheetnames) < 3:
        while len(wb.sheetnames) < 3:
            wb.create_sheet()
    if wb.sheetnames[2] != "Charts":
        charts_ws = wb.create_sheet("Charts")
        wb._sheets.insert(2, wb._sheets.pop())
    else:
        charts_ws = wb.worksheets[2]

    # Clear existing content
    for row in charts_ws.iter_rows():
        for cell in row:
            cell.value = None

    weekly_sheets = wb.sheetnames[3:]
    all_categories = {}
    model_data = {}
    category_totals = {}
    week_labels = []

    # Collect weekly data
    for sheetname in weekly_sheets:
        ws = wb[sheetname]
        week_labels.append(sheetname)

        last_category = None

        for row in range(2, ws.max_row + 1):
            category_cell = ws[f"A{row}"].value
            if category_cell and category_cell.strip():
                last_category = category_cell.strip()
                # Normalize similar category names
                if "cooler" in last_category.lower():
                    last_category = "Coolers"
                elif "chassis" in last_category.lower():
                    last_category = "Chassis"
                elif "power" in last_category.lower():
                    last_category = "Power Supplies"
            elif not last_category:
                continue  # No category yet

            category = last_category
            model = ws[f"B{row}"].value
            if not model or str(model).strip().upper().startswith("UPDATED"):
                continue

            # Keep order of models in category
            if category not in all_categories:
                all_categories[category] = []
            if model not in all_categories[category]:
                all_categories[category].append(model)

            # Sum columns C–AE
            total_value = 0
            for col in range(3, 32):  # C=3, AE=30
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, (int, float)):
                    total_value += cell_val

            # Store model data
            model_data.setdefault(category, {}).setdefault(model, []).append(total_value)

            # Store category totals
            category_totals.setdefault(category, [0] * len(weekly_sheets))
            category_totals[category][len(week_labels) - 1] += total_value

    # Write product data per category
    current_row = 1

    for category, models in all_categories.items():
        charts_ws.cell(row=current_row, column=1).value = f"{category} Stock Trends Data"
        start_row = current_row + 1
        charts_ws.cell(row=start_row, column=1).value = "Model"

        for i, week in enumerate(week_labels):
            charts_ws.cell(row=start_row, column=i + 2).value = week

        for j, model in enumerate(models):
            charts_ws.cell(row=start_row + j + 1, column=1).value = model
            model_values = model_data[category].get(model, [])
            for i, value in enumerate(model_values):
                charts_ws.cell(row=start_row + j + 1, column=i + 2).value = value

        # Create a table for this category
        end_row = start_row + len(models)
        end_col = 1 + len(week_labels)
        end_col_letter = get_column_letter(end_col)
        ref = f"A{start_row}:{end_col_letter}{end_row}"
        if end_row > start_row:
            add_or_update_table(charts_ws, f"{category.replace(' ', '')}Table", ref)

        current_row = start_row + len(models) + 3

    # Write category totals
    charts_ws.cell(row=current_row, column=1).value = "Category Totals Data"
    start_row = current_row + 1
    charts_ws.cell(row=start_row, column=1).value = "Category"
    # Preserve category order as first encountered
    ordered_categories = list(category_totals.keys())
    for i, week in enumerate(week_labels):
        charts_ws.cell(row=start_row, column=i + 2).value = week

    for w, cat in enumerate(ordered_categories):
        charts_ws.cell(row=start_row + w + 1, column=1).value = cat
        for i, week in enumerate(week_labels):
            charts_ws.cell(row=start_row + w + 1, column=i + 2).value = category_totals[cat][i]

    end_row = start_row + len(ordered_categories)
    end_col = 1 + len(week_labels)
    end_col_letter = get_column_letter(end_col)
    ref = f"A{start_row}:{end_col_letter}{end_row}"
    if end_row > start_row:
        add_or_update_table(charts_ws, "CategoryTotalsTable", ref)
    

    current_row = start_row + len(ordered_categories) + 3

    # Write store totals
    charts_ws.cell(row=current_row, column=1).value = "Store Totals Data"
    start_row = current_row + 1
    charts_ws.cell(row=start_row, column=1).value = "Store"
    for j, week in enumerate(week_labels):
        charts_ws.cell(row=start_row, column=j + 2).value = week

    # Each store becomes a row
    stores = list(store_map.values())
    for i, store in enumerate(stores):
        charts_ws.cell(row=start_row + i + 1, column=1).value = store
        for j, sheetname in enumerate(weekly_sheets):
            ws = wb[sheetname]
            store_total = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=3 + i).value
                if isinstance(val, (int, float)):
                    store_total += val
            charts_ws.cell(row=start_row + i + 1, column=j + 2).value = store_total

    end_row = start_row + len(stores)
    end_col = 1 + len(week_labels)
    end_col_letter = get_column_letter(end_col)
    ref = f"A{start_row}:{end_col_letter}{end_row}"
    if end_row > start_row:
        add_or_update_table(charts_ws, "StoreTotalsTable", ref)
    

def add_or_update_table(ws, base_name, ref):
    """Add a table if it doesn't exist, or update ref if it does, ensuring uniqueness."""

    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if max_row - min_row < 1:   # only header, no data
        return  # skip invalid table

    # Sanitize name
    safe_name = "".join(c if c.isalnum() else "_" for c in base_name)
    safe_name = safe_name[:200].lstrip("0123456789")  # Excel rules: no leading digit, < 255 chars

    # Collect existing tables
    if isinstance(ws._tables, dict):
        existing_tables = ws._tables
    else:
        existing_tables = {t.name: t for t in ws._tables}

    # Case 1: Update existing table with same name
    if safe_name in existing_tables:
        existing_tables[safe_name].ref = ref
        return

    # Case 3: Create new table with unique name
    tab = Table(displayName=safe_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def run_stock_tracker(target_wb, sheet_name):
    # Setup Selenium driver
    options = uc.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/120.0.0.0 Safari/537.36")
    driver = uc.Chrome(version_main=139, options=options)
    driver.maximize_window()

    # Setup worksheet
    ws = target_wb.create_sheet(title=sheet_name)
    headers = ["Product Category", "Model"] + list(store_map.values()) + ["INDIVIDUAL TOTALS", "CATEGORY TOTALS", "OUT OF STOCK"]
    ws.append(headers)

    # Start scanning for URLs
    for category, products in [("Power Supply", power_supplies),
                               ("Cooler", coolers),
                               ("Chassis", chassis),
                               ("Miscellaneous", miscellaneous)]:
        for name, url in products.items():
            if not url and DEBUG_MODE:
                print(f"Skipping {name}: URL is missing")
                continue
            row = [category, name]

            print(f"\nChecking stock for: {name}")

            for store_id, store_name in store_map.items():
                try:
                    stock = get_stock(url, store_id, driver)
                except Exception as e:
                    if DEBUG_MODE:
                        print(f"Error fetching stock for {name} at {store_name}: {str(e).splitlines()[0]}")
                    stock = 0
                print(f"{store_name}: {stock}")
                row.append(stock)
            ws.append(row)

    category_positions = format_new_sheet(ws)

    product_sums(ws, category_positions)

    driver.quit()
    
def get_stock(url, store_id, driver):
    try:
        driver.get("https://www.microcenter.com")
        driver.add_cookie({
            'name': 'storeSelected',
            'value': store_id,
            'domain': '.microcenter.com',
            'path': '/'
            })
        
        driver.get(url)
        human_delay(0.5, 1.5)

        try:
            stock_container = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "pnlInventory"))
            )
            if DEBUG_MODE:
                print("=== pnlInventory found ===")
        except TimeoutException:
            if DEBUG_MODE:
                print("pnlInventory not found, trying fallback .inventoryCnt")
            stock_container = None

        try:
            # Primary attempt: find OUT OF STOCK
            sold_out_element = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'SOLD OUT')]"))
            )
            if DEBUG_MODE:
                print("Found SOLD OUT element:", sold_out_element.get_attribute("outerHTML"))
            return 0  # Explicit out-of-stock
        except TimeoutException:
                # Second: check for inventory count
            try:
                stock_element = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "span.inventoryCnt"))
                )
                stock_text = stock_element.text.strip()
                if DEBUG_MODE:
                    print("Found inventoryCnt element:", stock_text)
            except TimeoutException:
                if DEBUG_MODE:
                    print("No stock element found, returning 0")
                return 0
            
        human_delay(0.5, 1.5)
        if DEBUG_MODE:
            print("stock_element found:", stock_element.get_attribute("outerHTML"))

        # Extract text content only
        stock_text = stock_element.text.strip()
        if DEBUG_MODE:
            print("Parsed stock_text:", stock_text)

        if "25+" in stock_text:
            return "25+"
        else:
            match = re.search(r"\d+", stock_text)
            return int(match.group(0)) if match else 0
    except (TimeoutException, NoSuchElementException):
        return 0
    except Exception as e:
        if DEBUG_MODE:
            print(f"get_stock failed for {url} at store {store_id}: {str(e).splitlines()[0]}")
        return 0
    
def human_delay(min_seconds=1.0, max_seconds=3.0):
    """Random delay to mimic human browsing."""
    delay = random.uniform(min_seconds, max_seconds)
    if DEBUG_MODE:
        print(f"Human-like delay for {delay:.2f} seconds")
    time.sleep(delay)

def terminate():
    sys.exit()


# Prompt user to add or remove products
def modify_products_window(use_original=False):
    filename = original_file if use_original else json_file

    win = tk.Toplevel()
    win.title("Modify Original Products" if use_original else "Modify Products")
    win.geometry("1200x700")
    win.protocol("WM_DELETE_WINDOW", terminate)

    with open(filename, "r") as f:
        data = json.load(f)
        local_power = data.get("power_supplies", {})
        local_coolers = data.get("coolers", {})
        local_chassis = data.get("chassis", {})
        local_misc = data.get("miscellaneous", {})

    def save_changes():
        with open(filename, "w") as f:
            json.dump({
                "power_supplies": local_power,
                "coolers": local_coolers,
                "chassis": local_chassis,
                "miscellaneous": local_misc
            }, f, indent=2)

    def refresh():
        if not win.winfo_exists():
            return
        
        all_items, idx = [], 1
        for category, products in [("Power Supplies", local_power),
                                   ("Coolers", local_coolers), 
                                   ("Chassis", local_chassis), 
                                   ("Miscellaneous", local_misc)]:
            for name, url in products.items():
                all_items.append((idx, category, name, url))
                idx += 1

        names_text.config(state="normal")
        names_text.delete("1.0", "end")
        names_text.insert("1.0", "\n".join(f"{i}. [{cat}] {name}" for i, cat, name, _ in all_items))
        names_text.config(state="disabled")

        urls_text.config(state="normal")
        urls_text.delete("1.0", "end")
        urls_text.insert("1.0", "\n".join(f"{i}. {url}" for i, _, _, url in all_items))
        urls_text.config(state="disabled")

        return all_items

    # Display model names
    tk.Label(win, text="Original Product Names" if use_original else "Product Names", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
    names_text = tk.Text(win, width=40, height=25, wrap="word")
    names_text.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
    names_text.config(state="disabled")

    # Display URLs
    tk.Label(win, text="Original Product URLs" if use_original else "Product URLs", font=("Arial", 12, "bold")).grid(row=0, column=1, padx=10, pady=5, sticky="w")
    urls_text = tk.Text(win, width=120, height=25, wrap="word")
    urls_text.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")
    urls_text.config(state="disabled")

    # Add and Remove buttons
    def add_product():
        add_win = tk.Toplevel(win)
        add_win.title("Add New Product")
        add_win.geometry("600x200")
        add_win.protocol("WM_DELETE_WINDOW", terminate)

        tk.Label(add_win, text="Model Name:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        model_entry = tk.Entry(add_win, width=50)
        model_entry.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(add_win, text="Product URL:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        url_entry = tk.Entry(add_win, width=50)
        url_entry.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(add_win, text="Category:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        category_var = tk.StringVar(value="Power Supply")
        tk.OptionMenu(add_win, category_var, "Power Supply", "Cooler", "Chassis", "None").grid(row=2, column=1, padx=10, pady=10, sticky="w")

        def confirm_add():
            new_model, new_url, category = model_entry.get().strip(), url_entry.get().strip(), category_var.get()

            if not new_model or not new_url:
                messagebox.showerror("Error", "Both Model Name and Product URL are required.")
                return
            if category == "Power Supply":
                local_power[new_model] = new_url
            elif category == "Cooler":
                local_coolers[new_model] = new_url
            elif category == "Chassis":
                local_chassis[new_model] = new_url
            else:
                local_misc[new_model] = new_url
            save_changes()
            messagebox.showinfo("Product Added", f"{new_model} added to {category}.")
            add_win.destroy()
            refresh()

        tk.Button(add_win, text="Add", command=confirm_add, width=15).grid(row=2, column=0, pady=20)
        tk.Button(add_win, text="Cancel", command=add_win.destroy, width=15).grid(row=2, column=1, pady=20)

    def remove_product():
        all_items = refresh()

        remove_model = simpledialog.askstring("Remove Product", "Enter the exact model name or model number to remove:")
        if not remove_model:
            return
        
        if remove_model.isdigit():
            remove_idx = int(remove_model)
            for idx, category, name, _ in all_items:
                if idx == remove_idx:
                    target_map = {
                        "Power Supplies": local_power,
                        "Coolers": local_coolers,
                        "Chassis": local_chassis,
                        "Miscellaneous": local_misc
                    }[category]
                    del target_map[name]
                    save_changes()
                    messagebox.showinfo("Product Removed", f"{name} removed.")
                    refresh()
                    return
        else:
            for product_map in [local_power, local_coolers, local_chassis, local_misc]:
                if remove_model in product_map:
                    del product_map[remove_model]
                    save_changes()
                    messagebox.showinfo("Product Removed", f"{remove_model} removed.")
                    refresh()
                    return

        messagebox.showinfo("Not Found", f"{remove_model} not found.")

    def reset_to_original():
        if not os.path.exists("MC_original_products.json"):
            messagebox.showerror("Error", "MC_original_products.json not found.")
            return
        with open(original_file, "r") as f:
            data = json.load(f)
            power_supplies.clear(); power_supplies.update(data.get("power_supplies", {}))
            coolers.clear(); coolers.update(data.get("coolers", {}))
            chassis.clear(); chassis.update(data.get("chassis", {}))
            miscellaneous.clear(); miscellaneous.update(data.get("miscellaneous", {}))
            # Also update the local GUI references
            local_power.clear(); local_power.update(power_supplies)
            local_coolers.clear(); local_coolers.update(coolers)
            local_chassis.clear(); local_chassis.update(chassis)
            local_misc.clear(); local_misc.update(miscellaneous)

        save_products(json_file)
        refresh()
        messagebox.showinfo("Reset Complete", "Product list reset to original.")


    def done():
        win.quit()
        win.destroy()

    tk.Button(win, text="Add Product", command=add_product, width=20).grid(row=2, column=0, pady=10)
    tk.Button(win, text="Remove Product", command=remove_product, width=20).grid(row=2, column=1, pady=10)
    tk.Button(win, text="Done", command=done, width=20).grid(row=3, column=0, pady=10)
    if not use_original:  
        tk.Button(win, text="Reset to Original", command=reset_to_original, width=20).grid(row=3, column=1, pady=10)
        tk.Button(win, text="Modify Original", command=lambda: modify_products_window(use_original=True), width=20).grid(row=4, column=1, pady=10)

    refresh()
    win.mainloop()

def main():
    load_products()
    final_message = "Error"

    # Prompt user for file save location
    root = Tk()
    root.withdraw()  # Hide the main window
    root.protocol("WM_DELETE_WINDOW", terminate)

    # Ask the user for the week number (default is 1)
    week_number = simpledialog.askstring("Week Number", "Enter the week number (e.g., 9, 10, 11, etc):")
    if not week_number:
        terminate()

    if messagebox.askyesno("Modify Products", "Do you want to view, add, or remove products?"):
        modify_products_window()

    use_existing = messagebox.askyesno("Stock Tracker", "Would you like to edit an existing Excel file?")

    file_path = None
    # If the user owns the stock tracker file
    if use_existing:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            messagebox.showinfo("Cancelled", "No file selected. Exiting.")
            time.sleep(2)
            terminate()
    else:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="Save new stock report as")
        if not file_path:
            messagebox.showinfo("Cancelled", "No save location selected.")
            time.sleep(2)
            terminate()

    analyze_only = messagebox.askyesno("Analyze Only", "Would you like to ONLY run the highlighting/labeling program?")

    global STOCK_TRACKER_START
    STOCK_TRACKER_START = time.time()

    if use_existing:
        # Run the stock tracker and coloring
        wb = load_workbook(file_path)
        sheet_name = f"WK{week_number}"
        process_add_products_sheet(wb)
        if not analyze_only:
            run_stock_tracker(wb, sheet_name)
        prepare_chart_data(wb)
    else:
        # If the user opts to create an independent sheet with this week's stock
        wb = Workbook()
        ws = wb.active
        sheet_name = f"WK{week_number}"
        ws.title = sheet_name
        run_stock_tracker(wb, sheet_name)

    analyze_stock(wb)
    wb.save(file_path)
    elapsed = time.time() - STOCK_TRACKER_START
    processed_time = f"{int(elapsed // 3600)} Hours, {int((elapsed % 3600) // 60)} Minutes, {int((elapsed % 3600) % 60)} Seconds"
    final_message = f"Stock tracking completed in {processed_time}"
    print(final_message)
    os.startfile(file_path)

    time.sleep(2)
    terminate()

if __name__ == "__main__":
    main()